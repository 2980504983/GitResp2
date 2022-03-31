import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

# 1 创建
# 实例化一个excel文件对象在内存中,创建一个表文件在内存中
wb = Workbook()
# 打开已存在的文件
# wb2 = load_workbook('文件名称.xlsx')


# 2 表操作
# 获取表(被聚焦的表)
sheet = wb.active
# 修改表名
# sheet.title = 'qyzw'


# 3 写入数据
# 方式一:
sheet["C5"] = '数据C表示第3列,5表示第5行'

# 方式二:
sheet.append([1, 2, 3])  # 附加行,从第一列开始附加(添加在最后一行数据的后面)

# 方式三:
sheet['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")

# 2 保存表
wb.save('文件名.xlsx')
